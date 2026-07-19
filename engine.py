import io
import pandas as pd
from collections import defaultdict
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model

def parse_slots(s_str):
    s_str = s_str.strip()
    if len(s_str) == 1 and s_str in ["월", "화", "수", "목", "금"]:
        return [(s_str, p) for p in range(1, 8)]
    elif len(s_str) >= 2:
        return [(s_str[0], int(s_str[1:]))]
    return []

def run_timetable_engine(uploaded_file, user_conditions):
    xls = pd.ExcelFile(uploaded_file)
    target_sheet = xls.sheet_names[0]
    for sheet in xls.sheet_names:
        if '국어' in str(sheet).replace(" ", ""):
            target_sheet = sheet
            break
            
    df_kor = pd.read_excel(xls, sheet_name=target_sheet, header=None)

    common_classes = []
    seen = set()
    current_teacher = None
    current_subj = None
    
    support_teachers = [x.strip() for x in user_conditions.get("support_text", "").split(",") if x.strip()]

    for i in range(4, len(df_kor)):
        subj = df_kor.iloc[i, 0]
        teacher = df_kor.iloc[i, 1]

        if pd.notna(teacher) and str(teacher).strip() not in ['담당교사', '담당', '교사', 'nan', '']:
            current_teacher = str(teacher).strip()
        if pd.notna(subj) and str(subj).strip() not in ['과목', '교과', 'nan', '']:
            current_subj = str(subj).strip()

        if not current_teacher or not current_subj: 
            continue

        if current_teacher in support_teachers:
            continue

        for c_idx in range(2, 26):
            val = df_kor.iloc[i, c_idx]
            try:
                val_float = float(val)
                if pd.notna(val) and val_float > 0:
                    if 2 <= c_idx <= 9: g, b = 1, c_idx - 1
                    elif 10 <= c_idx <= 17: g, b = 2, c_idx - 9
                    elif 18 <= c_idx <= 25: g, b = 3, c_idx - 17
                    key = (current_teacher, current_subj, g, b)
                    if key not in seen:
                        seen.add(key)
                        common_classes.append({
                            "teacher": current_teacher, "subj": current_subj,
                            "grade": g, "cls": b, "hours": int(val_float)
                        })
            except: pass

    teachers = sorted(list(set(c["teacher"] for c in common_classes)))
    classes = sorted(list(set((c["grade"], c["cls"]) for c in common_classes)))

    # 에러 방지: 요일별 시간표 슬롯을 7교시로 넉넉하게 오픈하여 시수 초과 크래시 원천 차단
    GRID = {"월":7, "화":7, "수":7, "목":7, "금":7}
    DAYS = list(GRID.keys())
    slots = [(d, p) for d in DAYS for p in range(1, GRID[d] + 1)]
    S = len(slots); sidx = {s: i for i, s in enumerate(slots)}

    base_hard_rules = list(user_conditions.get("hard_rules", []))
    base_soft_rules = list(user_conditions.get("soft_rules", []))
    
    # 우선순위: 3연강 절대 금지는 최후의 최후까지 지켜냅니다.
    relax_order = [
        "특정 학년-요일 금지 (무용 등)",
        "동일 학급 1일 1과목 분산",
        "1일 수업 시수 균등 배정",
        "4교시(점심) 공강 담임별 균등",
        "1교시 공강 1시간 필수",
        "미술 블록 오전/오후 균등",
        "운동장 체육 2학급 이하 제한",
        "교과 3연강 절대 금지"
    ]

    attempts = []
    current_hr = list(base_hard_rules)
    relaxed_so_far = []
    
    attempts.append((current_hr.copy(), True, [])) 
    
    for r in relax_order:
        if r in current_hr:
            current_hr.remove(r)
            relaxed_so_far.append(r)
            attempts.append((current_hr.copy(), True, relaxed_so_far.copy()))
            
    attempts.append((current_hr.copy(), False, relaxed_so_far.copy() + ["[비상배정] 일부 교사의 개인 금지/필수 시간 강제 침범"]))

    st = None
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 30 # 막히면 빠르게 다음 완화 단계로 넘어감

    final_xc = None
    final_out = None
    final_feedback = ""

    for attempt_hr, strict_bans, attempt_history in attempts:
        m = cp_model.CpModel()
        xc = {(ci, i): m.NewBoolVar(f"c{ci}_{i}") for ci in range(len(common_classes)) for i in range(S)}
        tocc = {}
        for t in teachers:
            cidx = [ci for ci, u in enumerate(common_classes) if u["teacher"] == t]
            for i in range(S):
                o = m.NewBoolVar("")
                m.Add(o == sum(xc[(ci, i)] for ci in cidx))
                tocc[(t, i)] = o
            for i in range(S): m.Add(tocc[(t, i)] <= 1)

        ban = defaultdict(set)
        reserved = set()
        for i in [sidx[(d, p)] for d, p in slots if d=="월" and p==7]:
            for g, b in classes: reserved.add((g, b, i)) 

        for line in user_conditions.get("banned_text", "").split("\n"):
            if ":" in line:
                t, slots_part = line.split(":", 1)
                t = t.strip()
                for s in slots_part.split(","):
                    for d, p in parse_slots(s):
                        if (d, p) in sidx: ban[t].add(sidx[(d, p)])

        if user_conditions.get("opt_bujang_free", True):
            bujang = [x.strip() for x in user_conditions.get("bujang_text", "").split(",") if x.strip()]
            for t in bujang:
                if ("금", 6) in sidx: ban[t].add(sidx[("금", 6)])

        fixed_slots = []
        sports_slot_indices = set()
        for line in user_conditions.get("pins_text", "").split("\n"):
            line = line.strip()
            if not line or line.startswith("["): continue
            for part in line.split("/"):
                part = part.strip()
                if "(" in part and "):" in part:
                    cls_part, rest = part.split("(", 1)
                    slot_part, teachers_part = rest.split("):", 1)
                    cls_part, slot_part = cls_part.strip(), slot_part.strip()
                    if len(slot_part) >= 2 and (slot_part[0], int(slot_part[1:])) in sidx:
                        slot_i = sidx[(slot_part[0], int(slot_part[1:]))]
                        if "-" in cls_part:
                            sports_slot_indices.add(slot_i)
                            g, b = map(int, cls_part.split("-"))
                            cls_tuple = (g, b)
                            subj_name = "스포츠"
                        else:
                            cls_tuple = None
                            subj_name = "지원"
                        for t in teachers_part.split(","):
                            t = t.strip()
                            ban[t].add(slot_i)
                            fixed_slots.append({"teacher": t, "cls": cls_tuple, "slot": slot_i, "subj": subj_name})

        def is_active(t, d, p):
            if (d, p) not in sidx: return 0
            slot_i = sidx[(d, p)]
            is_fixed = 1 if any(f["teacher"] == t and f["slot"] == slot_i for f in fixed_slots) else 0
            if t not in teachers: return is_fixed
            return tocc[(t, slot_i)] + is_fixed

        pen = []
        
        sp_teachers = [x.strip() for x in user_conditions.get("special_room_text", "").split(",") if x.strip()]
        if len(sp_teachers) >= 2:
            t1, t2 = sp_teachers[0], sp_teachers[1]
            if t1 in teachers and t2 in teachers:
                for i in range(S):
                    if strict_bans: m.Add(tocc[(t1, i)] + tocc[(t2, i)] <= 1)
                    else:
                        overlap = m.NewIntVar(0, 1, "")
                        m.Add(overlap >= tocc[(t1, i)] + tocc[(t2, i)] - 1)
                        pen.append((1000, overlap))

        for ci, u in enumerate(common_classes):
            g, b = u["grade"], u["cls"]; t = u["teacher"]
            m.Add(sum(xc[(ci, i)] for i in range(S)) == u["hours"])
            for i in range(S):
                if (g, b, i) in reserved: 
                    m.Add(xc[(ci, i)] == 0)
                elif i in ban[t]:
                    if strict_bans: 
                        m.Add(xc[(ci, i)] == 0)
                    else: 
                        pen.append((1000, xc[(ci, i)])) # 비상시 개인 금지시간 어기면 페널티

        for g, b in classes:
            cidx = [ci for ci, u in enumerate(common_classes) if u["grade"] == g and u["cls"] == b]
            for i in range(S):
                sports_in_class_slot = any(f["cls"] == (g, b) and f["slot"] == i for f in fixed_slots)
                if sports_in_class_slot:
                    m.Add(sum(xc[(ci, i)] for ci in cidx) == 0)
                else:
                    m.Add(sum(xc[(ci, i)] for ci in cidx) <= 1)

        for line in user_conditions.get("mandatory_text", "").split("\n"):
            if ":" in line and ">=" in line:
                left, req = line.rsplit(">=", 1)
                t, slots_part = left.split(":", 1)
                t, req = t.strip(), int(req.strip())
                target_slots = []
                for s in slots_part.split(","):
                    for d, p in parse_slots(s):
                        if (d, p) in sidx: target_slots.append(sidx[(d, p)])
                cidx_t = [ci for ci, u in enumerate(common_classes) if u["teacher"] == t]
                if cidx_t and target_slots:
                    expr = sum(xc[(ci, i)] for ci in cidx_t for i in target_slots)
                    if strict_bans:
                        m.Add(expr >= req)
                    else:
                        shortfall = m.NewIntVar(0, req, "")
                        m.Add(shortfall >= req - expr)
                        pen.append((1000, shortfall))

        # [수정완료] 블록타임제 3학년은 무조건 2시간 블록 1개 생성 (남은 1시간이 있다면 자동 분산)
        block_teachers = [x.strip() for x in user_conditions.get("block_text", "").split(",") if x.strip()]
        for ci, u in enumerate(common_classes):
            if u["teacher"] in block_teachers and u["grade"] == 3 and u["hours"] >= 2:
                block_vars = []
                for d in DAYS:
                    for p in range(1, GRID[d]):
                        if p == 4: continue
                        if (d, p) in sidx and (d, p+1) in sidx:
                            b_var = m.NewBoolVar(f"block_{ci}_{d}_{p}")
                            block_vars.append((b_var, sidx[(d, p)], sidx[(d, p+1)]))
                if block_vars:
                    m.AddExactlyOne([bv[0] for bv in block_vars])
                    for i in range(S):
                        is_in_block = sum(bv[0] for bv in block_vars if bv[1] == i or bv[2] == i)
                        m.Add(xc[(ci, i)] >= is_in_block)
            
            # 단, 블록타임제(2H)와 교과(1H)가 만나서 3연강이 되는 참사는 무조건 방지!
            if u["hours"] >= 3:
                for d in DAYS:
                    for p in range(1, GRID[d] - 1):
                        if all((d, p+k) in sidx for k in range(3)):
                            m.Add(sum(xc[(ci, sidx[(d, p+k)])] for k in range(3)) <= 2)

        # ==== 동적 자동 완화 적용 영역 ====
        if "특정 학년-요일 금지 (무용 등)" in attempt_hr or "특정 학년-요일 금지 (무용 등)" in base_soft_rules:
            for line in user_conditions.get("grade_day_text", "").split("\n"):
                if ":" in line:
                    t, rule_part = line.split(":", 1)
                    t = t.strip()
                    for rule in rule_part.split(","):
                        if "-" in rule:
                            g_str, day_str = rule.split("-", 1)
                            try:
                                g_val, day_str = int(g_str.strip()), day_str.strip()
                                for ci, u in enumerate(common_classes):
                                    if u["teacher"] == t and u["grade"] == g_val:
                                        target_slots = [sidx[(day_str, p)] for p in range(1, GRID[day_str] + 1) if (day_str, p) in sidx]
                                        if "특정 학년-요일 금지 (무용 등)" in attempt_hr:
                                            for i in target_slots: m.Add(xc[(ci, i)] == 0)
                                        else:
                                            for i in target_slots: pen.append((30, xc[(ci, i)]))
                            except: pass

        if "동일 학급 1일 1과목 분산" in attempt_hr or "동일 학급 1일 1과목 분산" in base_soft_rules:
            byc = defaultdict(list)
            for ci, u in enumerate(common_classes): byc[((u["grade"], u["cls"]), u["teacher"])].append(ci)
            for (cls, t), idxs in byc.items():
                if len(idxs) >= 1:
                    total_h = sum(common_classes[ci]["hours"] for ci in idxs)
                    
                    # 블록 교사일 경우 일일 최대 시수 한도를 똑똑하게 2로 올려주어 에러 방지
                    is_block = any(common_classes[ci]["teacher"] in block_teachers and common_classes[ci]["grade"] == 3 and common_classes[ci]["hours"] >= 2 for ci in idxs)
                    
                    limit = max(1, (total_h + 4) // 5) 
                    if is_block: limit = max(limit, 2)
                    
                    for d in DAYS:
                        target_slots = [sidx[(d, p)] for p in range(1, GRID[d] + 1) if (d, p) in sidx]
                        expr = sum(xc[(ci, i)] for ci in idxs for i in target_slots)
                        if "동일 학급 1일 1과목 분산" in attempt_hr:
                            m.Add(expr <= limit)
                        else:
                            excess = m.NewIntVar(0, 5, "")
                            m.Add(excess >= expr - limit)
                            pen.append((20, excess))

        if "교과 3연강 절대 금지" in attempt_hr or "교과 3연강 절대 금지" in base_soft_rules:
            for t in teachers:
                if t in support_teachers: continue 
                for d in DAYS:
                    for p in range(1, GRID[d] - 2):
                        if all((d, p+k) in sidx for k in range(4)):
                            expr_4 = sum(is_active(t, d, p+k) for k in range(4))
                            m.Add(expr_4 <= 3) # 어떤 타협에도 4연강은 절대 사수
                    for p in range(1, GRID[d] - 1):
                        if all((d, p+k) in sidx for k in range(3)):
                            expr_reg_3 = sum(tocc[(t, sidx[(d, p+k)])] for k in range(3))
                            if "교과 3연강 절대 금지" in attempt_hr:
                                m.Add(expr_reg_3 <= 2)
                            else:
                                excess = m.NewIntVar(0, 3, "")
                                m.Add(excess >= expr_reg_3 - 2)
                                pen.append((50, excess))

        if "1일 수업 시수 균등 배정" in attempt_hr or "1일 수업 시수 균등 배정" in base_soft_rules:
            for t in teachers:
                day_sums = [sum(is_active(t, d, p) for p in range(1, GRID[d]+1) if (d, p) in sidx) for d in DAYS]
                d_max = m.NewIntVar(0, 7, ""); d_min = m.NewIntVar(0, 7, "")
                m.AddMaxEquality(d_max, day_sums); m.AddMinEquality(d_min, day_sums)
                if "1일 수업 시수 균등 배정" in attempt_hr:
                    m.Add(d_max - d_min <= 2)
                else:
                    pen.append((10, d_max - d_min))

        # [수정완료] 1교시 공강 1시간을 강제하다 에러나는 현상 수학적 해결 (주당 최소 1회 이상 공강 보장)
        if "1교시 공강 1시간 필수" in attempt_hr or "1교시 공강 1시간 필수" in base_soft_rules:
            for t in teachers:
                s_1 = sum(is_active(t, d, 1) for d in DAYS if (d, 1) in sidx)
                if "1교시 공강 1시간 필수" in attempt_hr:
                    m.Add(s_1 <= 4)
                else:
                    excess = m.NewIntVar(0, 5, "")
                    m.Add(excess >= s_1 - 4)
                    pen.append((15, excess))

        if "4교시(점심) 공강 담임별 균등" in attempt_hr or "4교시(점심) 공강 담임별 균등" in base_soft_rules:
            hr_1 = [x.strip() for x in user_conditions.get("hr1_text", "").split(",") if x.strip()]
            hr_2 = [x.strip() for x in user_conditions.get("hr2_text", "").split(",") if x.strip()]
            hr_3 = [x.strip() for x in user_conditions.get("hr3_text", "").split(",") if x.strip()]
            hr_others = [t for t in teachers if t not in set(hr_1 + hr_2 + hr_3)]
            
            for g_list in [hr_1, hr_2, hr_3, hr_others]:
                valid_t = [t for t in g_list if t in teachers]
                if not valid_t: continue
                p4_sums = [sum(is_active(t, d, 4) for d in DAYS if (d, 4) in sidx) for t in valid_t]
                g_max = m.NewIntVar(0, 5, ""); g_min = m.NewIntVar(0, 5, "")
                m.AddMaxEquality(g_max, p4_sums); m.AddMinEquality(g_min, p4_sums)
                if "4교시(점심) 공강 담임별 균등" in attempt_hr:
                    m.Add(g_max - g_min <= 1)
                else:
                    pen.append((5, g_max - g_min))

        if "운동장 체육 2학급 이하 제한" in attempt_hr or "운동장 체육 2학급 이하 제한" in base_soft_rules:
            for i in sports_slot_indices:
                pe_in_slot = []
                for ci, u in enumerate(common_classes):
                    if "체육" in u["subj"] and "지원" not in u["subj"]:
                        pe_in_slot.append(xc[(ci, i)])
                if "운동장 체육 2학급 이하 제한" in attempt_hr:
                    m.Add(sum(pe_in_slot) <= 2)
                else:
                    excess = m.NewIntVar(0, 10, "")
                    m.Add(excess >= sum(pe_in_slot) - 2)
                    pen.append((50, excess))

        if "미술 블록 오전/오후 균등" in attempt_hr or "미술 블록 오전/오후 균등" in base_soft_rules:
            for t in block_teachers:
                cidx_3 = [ci for ci, u in enumerate(common_classes) if u["teacher"] == t and u["hours"] == 2]
                if cidx_3:
                    am = sum(xc[(ci, sidx[(d, p)])] for ci in cidx_3 for d in DAYS for p in range(1, 5) if (d, p) in sidx)
                    pm = sum(xc[(ci, sidx[(d, p)])] for ci in cidx_3 for d in DAYS for p in range(5, 8) if (d, p) in sidx)
                    diff = m.NewIntVar(-10, 10, "")
                    m.Add(diff == am - pm)
                    abs_diff = m.NewIntVar(0, 10, "")
                    m.AddAbsEquality(abs_diff, diff)
                    if "미술 블록 오전/오후 균등" in attempt_hr:
                        m.Add(abs_diff <= 1)
                    else:
                        pen.append((20, abs_diff))

        m.Minimize(sum(w * v for w, v in pen))
        st = solver.Solve(m)

        if st in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            final_xc = xc
            if attempt_history:
                final_feedback = "⚠️ [비상 타협 모드 작동]\n완벽한 조건을 찾지 못해, 엑셀 출력을 위해 다음 규칙들을 억지로 어겨가며 배정했습니다.\n\n어긴 항목: " + ", ".join(attempt_history)
            break 

    # ==== 엑셀 출력 ====
    if final_xc is not None:
        out = {}
        for f in fixed_slots:
            t_name = f["teacher"]
            if t_name not in out: out[t_name] = []
            cls_str = f"{f['cls'][0]}-{f['cls'][1]}" if f["cls"] else ""
            out[t_name].append({"subj": f["subj"], "cls": cls_str, "slot": f["slot"]})
            if t_name not in teachers: teachers.append(t_name)
        
        for ci, u in enumerate(common_classes):
            for i in range(S):
                if solver.Value(final_xc[(ci, i)]):
                    t_name = u["teacher"]
                    if t_name not in out: out[t_name] = []
                    out[t_name].append({"subj": u["subj"], "cls": f"{u['grade']}-{u['cls']}", "slot": i})
        
        custom_order = [x.strip() for x in user_conditions.get("sort_text", "").split(",") if x.strip()]
        group_ends = [x.strip() for x in user_conditions.get("border_text", "").split(",") if x.strip()]
        sorted_teachers = sorted(out.keys(), key=lambda x: custom_order.index(x) if x in custom_order else 999)
                    
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "교사별 시간표"
        
        # 에러 완전 차단을 위해 7교시까지 출력 (없는 교시는 빈칸으로 깔끔하게 처리됩니다)
        days_periods = [("월", 7), ("화", 7), ("수", 7), ("목", 7), ("금", 7)]

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_options.horizontalCentered = True

        font_title1 = Font(name="바탕체", size=14, bold=True)
        font_title2 = Font(name="바탕체", size=11, bold=True)
        font_base = Font(name="바탕체", size=8)
        font_bold = Font(name="바탕체", size=9, bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin = Side(style="thin", color="000000")
        thick = Side(style="medium", color="000000")
        bd_normal = Border(left=thin, right=thin, top=thin, bottom=thin)
        bd_thick_bottom = Border(left=thin, right=thin, top=thin, bottom=thick)

        ws.cell(1, 1, "2026-2학기 전체교사 시간표 초안").font = font_title1
        ws.cell(2, 1, "교사별 시간표").font = font_title2
        ws.cell(3, 1, "교사").font = font_bold
        ws.cell(3, 1).alignment = center; ws.cell(3, 1).border = bd_normal; ws.cell(4, 1, "").border = bd_normal
        ws.merge_cells("A3:A4")

        col = 2
        for day, periods in days_periods:
            ws.cell(3, col, day).font = font_bold
            ws.cell(3, col).alignment = center; ws.cell(3, col).border = bd_normal
            ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+periods-1)
            for p in range(1, periods + 1):
                c = ws.cell(4, col, str(p))
                c.font = font_bold; c.alignment = center; c.border = bd_normal
                ws.cell(3, col).border = bd_normal
                col += 1

        ws.cell(3, col, "교사").font = font_bold
        ws.cell(3, col).alignment = center; ws.cell(3, col).border = bd_normal; ws.cell(4, col, "").border = bd_normal
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)

        ws.cell(3, col+1, "계").font = font_bold
        ws.cell(3, col+1).alignment = center; ws.cell(3, col+1).border = bd_normal; ws.cell(4, col+1, "").border = bd_normal
        ws.merge_cells(start_row=3, start_column=col+1, end_row=4, end_column=col+1)
        total_cols = col + 1

        r = 5
        for t_name in sorted_teachers:
            items = out[t_name]
            slot_map = {item["slot"]: item for item in items}
            
            is_group_end = t_name in group_ends
            bottom_bd = bd_thick_bottom if is_group_end else bd_normal
            
            ws.cell(r, 1, t_name).font = font_base
            ws.cell(r, 1).alignment = center; ws.cell(r, 1).border = bd_normal; ws.cell(r+1, 1).border = bottom_bd
            ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=1)
            
            col = 2
            count = 0
            for i in range(sum([p for _, p in days_periods])):
                if i in slot_map:
                    cls_str = slot_map[i]["cls"]
                    subj_str = slot_map[i]["subj"]
                    count += 1
                else:
                    cls_str = ""
                    subj_str = ""
                    
                c1 = ws.cell(r, col, cls_str); c2 = ws.cell(r+1, col, subj_str)
                c1.font = font_base; c1.alignment = center; c1.border = bd_normal
                c2.font = font_base; c2.alignment = center; c2.border = bottom_bd
                col += 1
                
            ws.cell(r, col, t_name).font = font_base
            ws.cell(r, col).alignment = center; ws.cell(r, col).border = bd_normal; ws.cell(r+1, col).border = bottom_bd
            ws.merge_cells(start_row=r, start_column=col, end_row=r+1, end_column=col)
            
            ws.cell(r, col+1, count).font = font_base
            ws.cell(r, col+1).alignment = center; ws.cell(r, col+1).border = bd_normal; ws.cell(r+1, col+1).border = bottom_bd
            ws.merge_cells(start_row=r, start_column=col+1, end_row=r+1, end_column=col+1)
            r += 2

        ws.column_dimensions["A"].width = 5.0
        for i in range(2, total_cols - 1):
            ws.column_dimensions[get_column_letter(i)].width = 3.6
        ws.column_dimensions[get_column_letter(total_cols-1)].width = 5.0
        ws.column_dimensions[get_column_letter(total_cols)].width = 5.0

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output, "성공", final_feedback
        
    else:
        return None, "실패", "서버 응답 오류"